#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""读取和解析Excel文件内容"""

import sys
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def read_excel(file_path, max_rows=100, max_cols=20):
    """
    读取Excel文件并输出为JSON格式
    
    参数:
        file_path: Excel文件路径
        max_rows: 最大读取行数（默认100）
        max_cols: 最大读取列数（默认20）
    """
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        result = {
            "file": file_path,
            "sheets": []
        }
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 获取实际使用的行列范围
            max_row = min(sheet.max_row, max_rows)
            max_col = min(sheet.max_column, max_cols)
            
            sheet_data = {
                "name": sheet_name,
                "rows": max_row,
                "cols": max_col,
                "data": []
            }
            
            # 读取数据
            for row_idx, row in enumerate(sheet.iter_rows(max_row=max_row, max_col=max_col), 1):
                row_data = []
                for cell in row:
                    value = cell.value
                    # 转换为字符串，处理None值
                    if value is None:
                        row_data.append("")
                    else:
                        row_data.append(str(value))
                
                sheet_data["data"].append({
                    "row": row_idx,
                    "values": row_data
                })
            
            result["sheets"].append(sheet_data)
        
        return result
        
    except Exception as e:
        return {
            "error": str(e),
            "file": file_path
        }

def format_output(data, format_type="json"):
    """
    格式化输出
    
    参数:
        data: 读取的数据
        format_type: 输出格式 (json/markdown)
    """
    if "error" in data:
        return f"错误: {data['error']}"
    
    if format_type == "json":
        return json.dumps(data, ensure_ascii=False, indent=2)
    
    elif format_type == "markdown":
        output = [f"# Excel文件: {data['file']}\n"]
        
        for sheet in data["sheets"]:
            output.append(f"\n## 工作表: {sheet['name']}")
            output.append(f"行数: {sheet['rows']}, 列数: {sheet['cols']}\n")
            
            # 创建Markdown表格
            if len(sheet["data"]) > 0:
                # 表头（第一行）
                header = sheet["data"][0]["values"]
                output.append("| " + " | ".join(header) + " |")
                output.append("| " + " | ".join(["---"] * len(header)) + " |")
                
                # 数据行（最多显示20行）
                for row in sheet["data"][1:21]:
                    output.append("| " + " | ".join(row["values"]) + " |")
                
                if len(sheet["data"]) > 21:
                    output.append(f"\n... 还有 {len(sheet['data']) - 21} 行数据\n")
        
        return "\n".join(output)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python read_xlsx.py <excel文件路径> [输出格式:json|markdown]")
        sys.exit(1)
    
    file_path = sys.argv[1]
    format_type = sys.argv[2] if len(sys.argv) > 2 else "markdown"
    
    data = read_excel(file_path)
    output = format_output(data, format_type)
    print(output)
